import os
import csv
import xlrd
from openpyxl import load_workbook
from unidecode import unidecode

input_path = "agencias_xls"
output_path = "agencias_csv"

columns = [
    "cnpj",
    "sequencial do cnpj",
    "dv do cnpj",
    "nome instituicao",
    "nome da instituicao",
    "segmento",
    "segmentos",
    "cod compe ag",
    "nome da agencia",
    "nome agencia",
    "endereco",
    "bairro",
    "cep",
    "municipio",
    "estado",
    "uf",
]

for file in os.listdir(input_path):
    filename = os.path.join(input_path, file)

    year = file[0:4]
    month = file[4:6]

    if file.endswith(".xls"):
        workbook = xlrd.open_workbook(filename)
        worksheet = workbook.sheet_by_index(0)

        with open(
            os.path.join(output_path, f"{year}-{month}_agencias.csv"), "w", newline=""
        ) as csvfile:
            writer = csv.writer(csvfile, delimiter=",")
            permited_columns = []

            for row_num in range(0, worksheet.nrows):
                for col in range(worksheet.ncols):
                    cell_value = worksheet.cell_value(row_num, col)
                    if str(cell_value).strip().lower() == "cnpj":
                        header_number = row_num
                        print(header_number)
                        break

            for row_num in range(header_number, worksheet.nrows):
                row = []
                if row_num == header_number:
                    for col in range(worksheet.ncols):
                        cell_value = worksheet.cell_value(row_num, col)

                        if unidecode(str(cell_value).strip()).lower() in columns:
                            permited_columns.append(col)
                            if unidecode(str(cell_value).strip()).lower() == "estado":
                                cell_value = "uf"
                            if (
                                unidecode(str(cell_value).strip()).lower()
                                == "segmentos"
                            ):
                                cell_value = "segmento"
                            if unidecode(str(cell_value).strip()).lower() in [
                                "nome da agencia",
                                "nome agencia",
                            ]:
                                cell_value = "nome_agencia"
                            if unidecode(str(cell_value).strip()).lower() in [
                                "nome da instituicao",
                                "nome instituicao",
                            ]:
                                cell_value = "nome_instituicao"
                            cell_value = (
                                unidecode(str(cell_value).strip())
                                .lower()
                                .replace(" ", "_")
                            )
                            row.append(cell_value)
                else:
                    for col in permited_columns:
                        cell_value = worksheet.cell_value(row_num, col)
                        if isinstance(cell_value, float):
                            cell_value = "{:.0f}".format(cell_value).replace(" ", "0")
                        row.append(cell_value)
                row = [str(cell).strip() for cell in row]
                row = filter(None, row)
                writer.writerow(row)
    else:
        workbook = load_workbook(filename=filename)

        sheet = workbook.active

        with open(
            os.path.join(output_path, f"{year}-{month}_agencias.csv"),
            mode="w",
            newline="",
        ) as csv_file:
            writer = csv.writer(csv_file, delimiter=",")

            permited_columns = []

            for index, row in enumerate(sheet.iter_rows(min_row=0, values_only=True)):
                for cell in row:
                    if unidecode(str(cell).strip()).lower() == "cnpj":
                        header_number = index
                        print(header_number)
                        break

            for index, row in enumerate(
                sheet.iter_rows(min_row=header_number + 1, values_only=True)
            ):
                if index == 0:
                    row = [str(cell).strip() for cell in row]
                    for cell in row:
                        if unidecode(cell).strip().lower() in columns:
                            permited_columns.append(row.index(cell))
                row = [str(cell).strip() for cell in row]
                new_row = []
                for cell in row:
                    if row.index(cell) in permited_columns:
                        if unidecode(str(cell).strip()).lower() == "estado":
                            cell = "uf"
                        if unidecode(str(cell).strip()).lower() == "segmentos":
                            cell = "segmento"
                        if unidecode(str(cell).strip()).lower() in [
                            "nome da agencia",
                            "nome agencia",
                        ]:
                            cell = "nome_agencia"
                        if unidecode(str(cell).strip()).lower() in [
                            "nome da instituicao",
                            "nome instituicao",
                        ]:
                            cell = "nome_instituicao"
                        if unidecode(str(cell).strip()).lower() in columns:
                            cell = (
                                unidecode(str(cell).strip()).lower().replace(" ", "_")
                            )
                        new_row.append(cell)
                new_row = filter(None, new_row)
                writer.writerow(new_row)

    print(filename)
